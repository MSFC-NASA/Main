import openpyxl
from pathlib import Path

# Setting the path to the xlsx file:
xlsx_file = Path('Book1.xlsx')
print(xlsx_file)

wb_obj = openpyxl.load_workbook(xlsx_file)
print(wb_obj)

sheet = wb_obj.active
print(sheet)

print(sheet["C2"].value)
for row in sheet.iter_rows(max_row=6):
    for cell in row:
        print(cell.value, end=" ")
    print()

print(sheet.max_row, sheet.max_column)
